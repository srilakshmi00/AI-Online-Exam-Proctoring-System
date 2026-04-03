"""
Export Database to Excel
=========================
Project : Online Exam Proctoring System Based on AI
Authors : Ramkumar R, Sri Lakshmi T, Subiksha V
College : Shree Venkateshwara Hi-Tech Engineering College

Exports all 4 database tables into one formatted Excel file:
    Sheet 1 — Summary Dashboard
    Sheet 2 — Candidates
    Sheet 3 — Exam Sessions
    Sheet 4 — Event Logs (Violations)
    Sheet 5 — Results

INSTALL (run once):
    pip install openpyxl mysql-connector-python

RUN:
    python export_to_excel.py
"""

import mysql.connector
from openpyxl import Workbook
from openpyxl.styles import (Font, PatternFill, Alignment,
                               Border, Side, GradientFill)
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, Reference, PieChart
from datetime import datetime
import os

# ─────────────────────────────────────────────
#  !! UPDATE YOUR PASSWORD HERE !!
# ─────────────────────────────────────────────
DB_CONFIG = {
    "host":     "localhost",
    "user":     "root",
    "password": "password",
    "database": "proctoring_db"
}

OUTPUT_FILE = f"Proctoring_Report_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"

# ─────────────────────────────────────────────
#  COLOUR PALETTE
# ─────────────────────────────────────────────
C_HEADER_BG  = "1A1A2E"   # dark navy — header row background
C_HEADER_FG  = "FFFFFF"   # white — header text
C_TITLE_BG   = "E94560"   # red-pink — sheet title
C_TITLE_FG   = "FFFFFF"
C_ALT_ROW    = "F0F4FF"   # light blue-grey — alternating rows
C_HIGH       = "FF6B6B"   # red — HIGH severity
C_MEDIUM     = "FFB347"   # orange — MEDIUM severity
C_LOW        = "90EE90"   # green — LOW severity
C_PASS       = "90EE90"
C_FLAGGED    = "FFB347"
C_FAIL       = "FF6B6B"
C_BORDER     = "CCCCCC"
C_SUMMARY_BG = "16213E"
C_SUMMARY_FG = "E0E0FF"


# ═══════════════════════════════════════════════
#  HELPERS
# ═══════════════════════════════════════════════

def thin_border():
    s = Side(style="thin", color=C_BORDER)
    return Border(left=s, right=s, top=s, bottom=s)

def header_font(size=10):
    return Font(name="Arial", bold=True, color=C_HEADER_FG, size=size)

def body_font(size=10):
    return Font(name="Arial", size=size)

def header_fill():
    return PatternFill("solid", fgColor=C_HEADER_BG)

def alt_fill():
    return PatternFill("solid", fgColor=C_ALT_ROW)

def color_fill(hex_color):
    return PatternFill("solid", fgColor=hex_color)

def center():
    return Alignment(horizontal="center", vertical="center", wrap_text=True)

def left_align():
    return Alignment(horizontal="left", vertical="center", wrap_text=True)

def write_header_row(ws, row_num, columns, col_widths=None):
    """Write a styled header row."""
    for col_idx, col_name in enumerate(columns, start=1):
        cell = ws.cell(row=row_num, column=col_idx, value=col_name)
        cell.font      = header_font()
        cell.fill      = header_fill()
        cell.alignment = center()
        cell.border    = thin_border()
        if col_widths and col_idx - 1 < len(col_widths):
            ws.column_dimensions[get_column_letter(col_idx)].width = col_widths[col_idx - 1]

def write_title(ws, title, subtitle=""):
    """Write a coloured title banner at row 1."""
    ws.row_dimensions[1].height = 32
    cell = ws.cell(row=1, column=1, value=title)
    cell.font      = Font(name="Arial", bold=True, size=14, color=C_TITLE_FG)
    cell.fill      = color_fill(C_TITLE_BG)
    cell.alignment = center()
    if subtitle:
        ws.row_dimensions[2].height = 18
        sub = ws.cell(row=2, column=1, value=subtitle)
        sub.font      = Font(name="Arial", size=9, color="888888")
        sub.alignment = left_align()

def autofit_columns(ws, min_width=10, max_width=50):
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            try:
                if cell.value:
                    max_len = max(max_len, len(str(cell.value)))
            except Exception:
                pass
        ws.column_dimensions[col_letter].width = min(max(max_len + 3, min_width),
                                                      max_width)

def db_fetch(query, params=None):
    """Run a query and return (columns, rows)."""
    conn = mysql.connector.connect(**DB_CONFIG)
    cur  = conn.cursor()
    cur.execute(query, params or ())
    rows    = cur.fetchall()
    columns = [d[0] for d in cur.description]
    cur.close()
    conn.close()
    return columns, rows


# ═══════════════════════════════════════════════
#  SHEET 1 — SUMMARY DASHBOARD
# ═══════════════════════════════════════════════

def build_summary(wb):
    ws = wb.active
    ws.title = "Summary"
    ws.sheet_view.showGridLines = False
    ws.row_dimensions[1].height = 40

    # Title banner
    ws.merge_cells("A1:F1")
    t = ws["A1"]
    t.value     = "AI Online Exam Proctoring System — Report"
    t.font      = Font(name="Arial", bold=True, size=16, color=C_TITLE_FG)
    t.fill      = color_fill(C_TITLE_BG)
    t.alignment = center()

    ws.merge_cells("A2:F2")
    sub = ws["A2"]
    sub.value     = f"Generated: {datetime.now().strftime('%d %B %Y  %H:%M:%S')}  |  Shree Venkateshwara Hi-Tech Engineering College"
    sub.font      = Font(name="Arial", size=9, color="888888")
    sub.alignment = center()

    # Fetch counts
    _, rows = db_fetch("SELECT COUNT(*) FROM candidates")
    n_candidates = rows[0][0]

    _, rows = db_fetch("SELECT COUNT(*) FROM exam_sessions")
    n_sessions = rows[0][0]

    _, rows = db_fetch("SELECT COUNT(*) FROM event_logs")
    n_events = rows[0][0]

    _, rows = db_fetch("SELECT COUNT(*) FROM event_logs WHERE severity='HIGH'")
    n_high = rows[0][0]

    _, rows = db_fetch("SELECT COUNT(*) FROM results WHERE verdict='PASS'")
    n_pass = rows[0][0]

    _, rows = db_fetch("SELECT COUNT(*) FROM results WHERE verdict='FLAGGED'")
    n_flagged = rows[0][0]

    _, rows = db_fetch("SELECT COUNT(*) FROM results WHERE verdict='FAIL'")
    n_fail = rows[0][0]

    # Stat cards (row 4–7)
    stats = [
        ("A", "Enrolled Candidates", n_candidates, "1F4E79"),
        ("B", "Exam Sessions",       n_sessions,   "375623"),
        ("C", "Total Violations",    n_events,      "843C0C"),
        ("D", "High Severity",       n_high,        "9C0006"),
        ("E", "PASS",                n_pass,        "375623"),
        ("F", "FLAGGED / FAIL",      n_flagged + n_fail, "843C0C"),
    ]

    ws.row_dimensions[4].height = 20
    ws.row_dimensions[5].height = 36
    ws.row_dimensions[6].height = 16

    for col_letter, label, value, bg in stats:
        label_cell = ws[f"{col_letter}4"]
        label_cell.value     = label
        label_cell.font      = Font(name="Arial", size=9,
                                     bold=True, color="FFFFFF")
        label_cell.fill      = color_fill(bg)
        label_cell.alignment = center()

        val_cell = ws[f"{col_letter}5"]
        val_cell.value     = value
        val_cell.font      = Font(name="Arial", size=22,
                                   bold=True, color="FFFFFF")
        val_cell.fill      = color_fill(bg)
        val_cell.alignment = center()

    # Violation breakdown table (row 9+)
    _, viol_rows = db_fetch("""
        SELECT event_type, COUNT(*) as cnt,
               SUM(severity='HIGH') as high_cnt
        FROM event_logs
        GROUP BY event_type
        ORDER BY cnt DESC
    """)

    ws["A8"].value = "Violation Breakdown"
    ws["A8"].font  = Font(name="Arial", bold=True, size=11)

    headers = ["Violation Type", "Total Count", "High Severity Count"]
    write_header_row(ws, 9, headers, [28, 16, 22])

    for i, (etype, cnt, high_cnt) in enumerate(viol_rows, start=10):
        ws.cell(row=i, column=1, value=etype).font   = body_font()
        ws.cell(row=i, column=2, value=cnt).alignment = center()
        ws.cell(row=i, column=3, value=int(high_cnt or 0)).alignment = center()
        if i % 2 == 0:
            for c in range(1, 4):
                ws.cell(row=i, column=c).fill = alt_fill()
        for c in range(1, 4):
            ws.cell(row=i, column=c).border = thin_border()

    # Bar chart for violation types
    if viol_rows:
        chart      = BarChart()
        chart.type = "col"
        chart.title        = "Violations by Type"
        chart.y_axis.title = "Count"
        chart.style        = 10
        chart.height       = 12
        chart.width        = 20

        data_ref  = Reference(ws, min_col=2, max_col=2,
                              min_row=9, max_row=9 + len(viol_rows))
        cats_ref  = Reference(ws, min_col=1,
                              min_row=10, max_row=9 + len(viol_rows))
        chart.add_data(data_ref, titles_from_data=True)
        chart.set_categories(cats_ref)
        ws.add_chart(chart, "A" + str(11 + len(viol_rows)))

    for col in "ABCDEF":
        ws.column_dimensions[col].width = 22


# ═══════════════════════════════════════════════
#  SHEET 2 — CANDIDATES
# ═══════════════════════════════════════════════

def build_candidates(wb):
    ws = wb.create_sheet("Candidates")
    ws.sheet_view.showGridLines = False

    write_title(ws, "Enrolled Candidates",
                "All students registered in the proctoring system")

    columns, rows = db_fetch(
        "SELECT candidate_id, name, enrolled_at FROM candidates ORDER BY enrolled_at DESC"
    )
    headers    = ["Candidate ID", "Full Name", "Enrolled At"]
    col_widths = [20, 32, 24]
    write_header_row(ws, 4, headers, col_widths)

    ws.merge_cells("A1:C1")
    ws.merge_cells("A2:C2")

    for i, row in enumerate(rows, start=5):
        for j, val in enumerate(row, start=1):
            c = ws.cell(row=i, column=j,
                        value=str(val) if val else "")
            c.font      = body_font()
            c.alignment = center() if j != 2 else left_align()
            c.border    = thin_border()
            if i % 2 == 0:
                c.fill = alt_fill()

    ws.row_dimensions[4].height = 22
    ws["A3"].value = f"Total enrolled: {len(rows)}"
    ws["A3"].font  = Font(name="Arial", size=9, color="888888")


# ═══════════════════════════════════════════════
#  SHEET 3 — EXAM SESSIONS
# ═══════════════════════════════════════════════

def build_sessions(wb):
    ws = wb.create_sheet("Exam Sessions")
    ws.sheet_view.showGridLines = False

    write_title(ws, "Exam Sessions",
                "All exam sessions with duration and violation counts")

    _, rows = db_fetch("""
        SELECT s.session_id, s.candidate_id, c.name,
               s.exam_name, s.started_at, s.ended_at,
               s.violation_count, s.status
        FROM exam_sessions s
        LEFT JOIN candidates c ON c.candidate_id = s.candidate_id
        ORDER BY s.started_at DESC
    """)

    ws.merge_cells("A1:H1")
    ws.merge_cells("A2:H2")

    headers    = ["Session ID", "Candidate ID", "Name", "Exam",
                  "Started At", "Ended At", "Violations", "Status"]
    col_widths = [28, 16, 24, 22, 22, 22, 12, 14]
    write_header_row(ws, 4, headers, col_widths)

    STATUS_COLORS = {
        "active":    "D6E4F7",
        "completed": "D6F5D6",
        "flagged":   "FFD6D6"
    }

    for i, row in enumerate(rows, start=5):
        sid, cid, name, exam, started, ended, viols, status = row
        values = [sid, cid, name or "", exam,
                  str(started) if started else "",
                  str(ended)   if ended   else "In progress",
                  viols, status.upper() if status else ""]

        bg = STATUS_COLORS.get(status, "FFFFFF")

        for j, val in enumerate(values, start=1):
            c = ws.cell(row=i, column=j, value=val)
            c.font      = body_font()
            c.alignment = center() if j not in (3, 4) else left_align()
            c.border    = thin_border()
            c.fill      = color_fill(bg) if j == 8 else (
                alt_fill() if i % 2 == 0 else PatternFill())

    ws.row_dimensions[4].height = 22
    ws["A3"].value = f"Total sessions: {len(rows)}"
    ws["A3"].font  = Font(name="Arial", size=9, color="888888")


# ═══════════════════════════════════════════════
#  SHEET 4 — EVENT LOGS
# ═══════════════════════════════════════════════

def build_event_logs(wb):
    ws = wb.create_sheet("Event Logs")
    ws.sheet_view.showGridLines = False

    write_title(ws, "Violation Event Logs",
                "Every violation detected during exam sessions")

    _, rows = db_fetch("""
        SELECT e.logged_at, e.candidate_id, c.name,
               e.session_id, e.event_type,
               e.severity, e.description, e.screenshot_path
        FROM event_logs e
        LEFT JOIN candidates c ON c.candidate_id = e.candidate_id
        ORDER BY e.logged_at DESC
    """)

    ws.merge_cells("A1:H1")
    ws.merge_cells("A2:H2")

    headers    = ["Timestamp", "Candidate ID", "Name",
                  "Session ID", "Event Type",
                  "Severity", "Description", "Screenshot"]
    col_widths = [22, 16, 22, 28, 20, 12, 40, 30]
    write_header_row(ws, 4, headers, col_widths)

    SEV_COLORS = {
        "HIGH":   C_HIGH,
        "MEDIUM": C_MEDIUM,
        "LOW":    C_LOW
    }

    for i, row in enumerate(rows, start=5):
        logged, cid, name, sid, etype, sev, desc, screenshot = row
        values = [
            str(logged) if logged else "",
            cid, name or "", sid, etype,
            sev, desc or "",
            os.path.basename(screenshot) if screenshot else ""
        ]
        sev_color = SEV_COLORS.get(sev, "FFFFFF")

        for j, val in enumerate(values, start=1):
            c = ws.cell(row=i, column=j, value=val)
            c.font      = body_font()
            c.alignment = center() if j in (1, 2, 6) else left_align()
            c.border    = thin_border()
            # Colour severity cell
            if j == 6:
                c.fill = color_fill(sev_color)
                c.font = Font(name="Arial", size=10, bold=True)
            elif i % 2 == 0:
                c.fill = alt_fill()

    ws.row_dimensions[4].height = 22
    ws["A3"].value = f"Total events: {len(rows)}"
    ws["A3"].font  = Font(name="Arial", size=9, color="888888")

    # Freeze header rows
    ws.freeze_panes = "A5"


# ═══════════════════════════════════════════════
#  SHEET 5 — RESULTS
# ═══════════════════════════════════════════════

def build_results(wb):
    ws = wb.create_sheet("Results")
    ws.sheet_view.showGridLines = False

    write_title(ws, "Exam Results & Verdicts",
                "Final verdict for each completed session")

    _, rows = db_fetch("""
        SELECT r.candidate_id, c.name, r.session_id,
               r.total_violations, r.high_severity,
               r.medium_severity, r.verdict, r.generated_at
        FROM results r
        LEFT JOIN candidates c ON c.candidate_id = r.candidate_id
        ORDER BY r.generated_at DESC
    """)

    ws.merge_cells("A1:H1")
    ws.merge_cells("A2:H2")

    headers    = ["Candidate ID", "Name", "Session ID",
                  "Total Violations", "High Severity",
                  "Medium Severity", "Verdict", "Generated At"]
    col_widths = [16, 24, 28, 18, 15, 17, 12, 22]
    write_header_row(ws, 4, headers, col_widths)

    VERDICT_COLORS = {
        "PASS":    C_PASS,
        "FLAGGED": C_FLAGGED,
        "FAIL":    C_FAIL
    }

    for i, row in enumerate(rows, start=5):
        cid, name, sid, total, high, medium, verdict, gen_at = row
        values = [cid, name or "", sid, total, int(high or 0),
                  int(medium or 0), verdict,
                  str(gen_at) if gen_at else ""]

        v_color = VERDICT_COLORS.get(verdict, "FFFFFF")

        for j, val in enumerate(values, start=1):
            c = ws.cell(row=i, column=j, value=val)
            c.font      = body_font()
            c.alignment = center() if j != 2 else left_align()
            c.border    = thin_border()
            if j == 7:   # Verdict column
                c.fill = color_fill(v_color)
                c.font = Font(name="Arial", size=10, bold=True)
            elif i % 2 == 0:
                c.fill = alt_fill()

    ws.row_dimensions[4].height = 22
    ws["A3"].value = f"Total results: {len(rows)}"
    ws["A3"].font  = Font(name="Arial", size=9, color="888888")

    # Pie chart for verdicts
    if rows:
        _, verdict_counts = db_fetch("""
            SELECT verdict, COUNT(*) FROM results GROUP BY verdict
        """)
        if verdict_counts:
            pie_ws_row = 6 + len(rows)
            ws.cell(row=pie_ws_row, column=1, value="Verdict").font = header_font()
            ws.cell(row=pie_ws_row, column=2, value="Count").font   = header_font()
            ws.cell(row=pie_ws_row, column=1).fill = header_fill()
            ws.cell(row=pie_ws_row, column=2).fill = header_fill()

            for k, (v, cnt) in enumerate(verdict_counts, start=pie_ws_row + 1):
                ws.cell(row=k, column=1, value=v)
                ws.cell(row=k, column=2, value=cnt)

            pie = PieChart()
            pie.title  = "Verdict Distribution"
            pie.style  = 10
            pie.height = 12
            pie.width  = 16
            data = Reference(ws, min_col=2, max_col=2,
                             min_row=pie_ws_row,
                             max_row=pie_ws_row + len(verdict_counts))
            cats = Reference(ws, min_col=1,
                             min_row=pie_ws_row + 1,
                             max_row=pie_ws_row + len(verdict_counts))
            pie.add_data(data, titles_from_data=True)
            pie.dataLabels        = None
            pie.set_categories(cats)
            ws.add_chart(pie, f"D{pie_ws_row}")


# ═══════════════════════════════════════════════
#  MAIN EXPORT
# ═══════════════════════════════════════════════

def export():
    print("\n" + "=" * 52)
    print("  DB → Excel Export — AI Proctoring System")
    print("=" * 52)

    # Test connection
    try:
        conn = mysql.connector.connect(**DB_CONFIG)
        conn.close()
        print("  [OK] Connected to MySQL.")
    except mysql.connector.Error as e:
        print(f"\n  [ERROR] Cannot connect: {e}")
        print("  Fix DB_PASSWORD in this script and retry.\n")
        return

    wb = Workbook()

    print("  Building Summary sheet ...")
    build_summary(wb)

    print("  Building Candidates sheet ...")
    build_candidates(wb)

    print("  Building Exam Sessions sheet ...")
    build_sessions(wb)

    print("  Building Event Logs sheet ...")
    build_event_logs(wb)

    print("  Building Results sheet ...")
    build_results(wb)

    # Tab colours
    TAB_COLORS = {
        "Summary":       "E94560",
        "Candidates":    "1A6B8A",
        "Exam Sessions": "2E7D32",
        "Event Logs":    "8D1A0F",
        "Results":       "4A148C",
    }
    for sheet_name, color in TAB_COLORS.items():
        if sheet_name in wb.sheetnames:
            wb[sheet_name].sheet_properties.tabColor = color

    # Active sheet = Summary
    wb.active = wb["Summary"]

    wb.save(OUTPUT_FILE)
    print(f"\n  [SUCCESS] Saved: {OUTPUT_FILE}")
    print(f"  Location : {os.path.abspath(OUTPUT_FILE)}\n")


if __name__ == "__main__":
    export()
